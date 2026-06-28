# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 MDfier contributors
# https://github.com/sameer24688-jpg/MDfier
"""Conversion logic for MDfier.

Two directions:
  * to Markdown   - PDF/DOCX/XLSX/CSV/PPTX/HTML/images (incl. scanned docs via OCR)
  * from Markdown - DOCX / HTML / PDF / TXT / XLSX / CSV

All functions return a list of output file paths. Heavy imports are done lazily.
"""

from __future__ import annotations

import csv
import os
import re
import sys
from html import unescape
from html.parser import HTMLParser
from typing import Callable, List, Optional

StatusCb = Optional[Callable[[str], None]]

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif", ".webp"}
# Extensions routed to the "to Markdown" direction.
# NOTE: Excel (.xlsx/.xls) and CSV are intentionally excluded as INPUT - see the
# "Limitations" section in README.md. (Markdown -> .xlsx/.csv output is still supported.)
TO_MARKDOWN_EXTS = {
    ".pdf", ".docx", ".pptx", ".ppt",
    ".html", ".htm", ".json", ".xml", ".txt", ".epub", ".rtf",
} | IMAGE_EXTS
# Spreadsheet inputs are deliberately unsupported (lossy as Markdown). CSV is
# already model-readable plain text; Excel can be exported to CSV by the user.
EXCLUDED_INPUT_EXTS = {".xlsx", ".xls", ".csv"}
REVERSE_TARGETS = ("docx", "html", "pdf", "txt", "xlsx", "csv")

_MARKITDOWN = None


def resource_path(rel: str) -> str:
    """Resolve a bundled resource path (works under PyInstaller and from source)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def _get_markitdown():
    global _MARKITDOWN
    if _MARKITDOWN is None:
        from markitdown import MarkItDown

        _MARKITDOWN = MarkItDown()
    return _MARKITDOWN


class Cancelled(Exception):
    """Raised at a checkpoint when the caller requested cancellation."""


def _check_cancel(cancel_cb) -> None:
    if cancel_cb is not None and cancel_cb():
        raise Cancelled()


def _emit(status_cb: StatusCb, message: str) -> None:
    if status_cb:
        status_cb(message)


def _unique_path(path: str) -> str:
    """Return `path` if free, else a non-destructive 'name (1).ext' variant.

    Works for files and directories (a path with no extension just gets the
    counter appended, e.g. 'doc_images' -> 'doc_images (1)').
    """
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    counter = 1
    while True:
        candidate = f"{base} ({counter}){ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def _write_text(path: str, text: str) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(text)


def _read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as handle:
        return handle.read()


# --------------------------------------------------------------------------- #
# To Markdown
# --------------------------------------------------------------------------- #
PDF_PAGE_TEXT_THRESHOLD = 15


def image_to_markdown(path: str, status_cb: StatusCb = None, lang: str = "English") -> List[str]:
    import ocr

    _emit(status_cb, "Running local image OCR...")
    text = ocr.ocr_image_file(path, lang)
    out = _unique_path(os.path.splitext(path)[0] + ".md")
    header = f"# OCR Extraction from {os.path.basename(path)}\n\n"
    _write_text(out, header + (text or "_No text detected._"))
    return [out]


def _retag_figures(text: str, images_dirname: str, page_no: int) -> str:
    """Rewrite bare image refs to use a relative path and a placeholder alt-text."""
    def repl(match: "re.Match") -> str:
        filename = os.path.basename(match.group(1).strip())
        rel = f"{images_dirname}/{filename}"
        return f"![Figure (page {page_no}): {filename}]({rel})"

    return re.sub(r"!\[\]\(([^)]+)\)", repl, text)


def _as_blockquote(text: str, header: str) -> str:
    """Wrap OCR text in a Markdown blockquote so it reads as secondary context.

    Isolating image-derived text from the main body prevents minor OCR artifacts
    from polluting paragraph flow and signals downstream LLMs/RAG that this is
    graphic-extracted content.
    """
    lines = [f"> {header}", ">"]
    for line in text.splitlines():
        stripped = line.rstrip()
        lines.append(f"> {stripped}" if stripped else ">")
    return "\n".join(lines)


def _strip_boundary_overlap(prev: str, nxt: str) -> str:
    """Drop an exact duplicate line repeated across a page boundary.

    Conservative on purpose: only an exact match (after strip, length >= 8) of the
    last non-empty line of `prev` and the first non-empty line of `nxt` is removed,
    so legitimate repeated content is left untouched.
    """
    prev_lines = [ln for ln in prev.splitlines() if ln.strip()]
    if not prev_lines:
        return nxt
    tail = prev_lines[-1].strip()
    if len(tail) < 8:
        return nxt

    nxt_lines = nxt.splitlines()
    for index, line in enumerate(nxt_lines):
        if not line.strip():
            continue
        if line.strip() == tail:
            del nxt_lines[index]
        break
    return "\n".join(nxt_lines)


def pdf_to_markdown(path: str, status_cb: StatusCb = None, lang: str = "English",
                    cancel_cb=None) -> List[str]:
    """Layout-aware hybrid PDF -> Markdown.

    Digital pages go through pymupdf4llm (multi-column reading order, GFM tables,
    figure extraction). Image-only/scanned pages are OCR'd with RapidOCR using the
    selected language. Pages are merged in original order.
    """
    import fitz  # PyMuPDF
    import pymupdf4llm
    import ocr

    out = _unique_path(os.path.splitext(path)[0] + ".md")
    base = os.path.splitext(out)[0]
    images_dir = _unique_path(base + "_images")
    images_dirname = os.path.basename(images_dir)

    _emit(status_cb, "Analyzing PDF layout...")
    with fitz.open(path) as doc:
        total = doc.page_count
        digital = [i for i, page in enumerate(doc)
                   if len((page.get_text("text") or "").strip()) >= PDF_PAGE_TEXT_THRESHOLD]
    scanned = [i for i in range(total) if i not in set(digital)]

    page_md: dict = {}

    if digital:
        _emit(status_cb, "Extracting layout, tables, and figures...")
        os.makedirs(images_dir, exist_ok=True)
        chunks = pymupdf4llm.to_markdown(
            path,
            pages=digital,
            page_chunks=True,
            write_images=True,
            image_path=images_dir,
            image_format="png",
            dpi=150,
            # We classify pages and run our own language-aware RapidOCR on scanned
            # pages. Letting pymupdf4llm's layout engine also OCR digital pages makes
            # it emit the text layer AND an OCR of the same region, doubling/echoing
            # text (e.g. "Pd(OAc)2/dppb/dppb"). force_text keeps the real text layer.
            use_ocr=False,
            force_text=True,
        )
        for chunk in chunks:
            idx = chunk.get("metadata", {}).get("page_number", 0) - 1
            page_md[idx] = _retag_figures(chunk.get("text", ""), images_dirname, idx + 1)

    if scanned:
        with fitz.open(path) as doc:
            for count, i in enumerate(scanned, start=1):
                _check_cancel(cancel_cb)
                _emit(status_cb, f"OCR scanned page {i + 1} ({count}/{len(scanned)})...")
                image = ocr.render_pdf_page(doc[i])
                try:
                    text = ocr.ocr_pil_image(image, lang)
                finally:
                    image.close()
                body = text.strip() or "_No text detected._"
                page_md[i] = _as_blockquote(body, f"**OCR text (scanned page {i + 1}):**")

    ordered = [page_md.get(i, "").strip() for i in range(total)]
    parts: List[str] = []
    for chunk in ordered:
        if not chunk:
            continue
        if parts:
            chunk = _strip_boundary_overlap(parts[-1], chunk).strip()
        if chunk:
            parts.append(chunk)
    _write_text(out, "\n\n".join(parts))

    outputs = [out]
    if os.path.isdir(images_dir):
        if os.listdir(images_dir):
            outputs.append(images_dir)
        else:
            os.rmdir(images_dir)
    return outputs


def generic_to_markdown(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Converting to Markdown...")
    result = _get_markitdown().convert(path)
    out = _unique_path(os.path.splitext(path)[0] + ".md")
    _write_text(out, result.text_content or "")
    return [out]


def to_markdown(path: str, status_cb: StatusCb = None, lang: str = "English",
                cancel_cb=None) -> List[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext in EXCLUDED_INPUT_EXTS:
        raise ValueError(
            "Excel/CSV input is not supported (see README 'Limitations'). "
            "CSV is already plain text AI models read directly; export Excel to CSV."
        )
    _check_cancel(cancel_cb)
    if ext in IMAGE_EXTS:
        return image_to_markdown(path, status_cb, lang)
    if ext == ".pdf":
        return pdf_to_markdown(path, status_cb, lang, cancel_cb)
    return generic_to_markdown(path, status_cb)


# --------------------------------------------------------------------------- #
# Markdown helpers
# --------------------------------------------------------------------------- #
def _md_to_html_fragment(md_text: str) -> str:
    import markdown as md_lib

    return md_lib.markdown(
        md_text,
        extensions=["tables", "fenced_code", "sane_lists", "nl2br"],
    )


class _TextExtractor(HTMLParser):
    BLOCK_TAGS = {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6", "pre"}

    def __init__(self) -> None:
        super().__init__()
        self._parts: List[str] = []

    def handle_starttag(self, tag, attrs):
        if tag in self.BLOCK_TAGS:
            self._parts.append("\n")

    def handle_endtag(self, tag):
        if tag in self.BLOCK_TAGS:
            self._parts.append("\n")

    def handle_data(self, data):
        self._parts.append(data)

    def text(self) -> str:
        raw = "".join(self._parts)
        lines = [line.strip() for line in raw.splitlines()]
        cleaned: List[str] = []
        blank = False
        for line in lines:
            if line:
                cleaned.append(line)
                blank = False
            elif not blank:
                cleaned.append("")
                blank = True
        return "\n".join(cleaned).strip() + "\n"


def _strip_inline(text: str) -> str:
    """Reduce inline Markdown to readable plain text (for table cells)."""
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)   # images -> alt
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)    # links -> text
    text = re.sub(r"`([^`]*)`", r"\1", text)                # inline code
    text = re.sub(r"(\*\*|__)(.*?)\1", r"\2", text)         # bold
    text = re.sub(r"(\*|_)(.*?)\1", r"\2", text)            # italic
    return text.strip()


def extract_markdown_tables(md_text: str) -> List[List[List[str]]]:
    """Extract GFM pipe tables. Returns a list of tables; each table is a list of
    rows; each row a list of cell strings (header is the first row)."""
    lines = md_text.splitlines()
    n = len(lines)

    def split_row(line: str) -> List[str]:
        s = line.strip()
        if s.startswith("|"):
            s = s[1:]
        if s.endswith("|"):
            s = s[:-1]
        cells = re.split(r"(?<!\\)\|", s)
        return [_strip_inline(c.strip().replace("\\|", "|")) for c in cells]

    def is_separator(line: str) -> bool:
        s = line.strip()
        if "-" not in s or "|" not in (s + "|"):
            return False
        for cell in split_row(s):
            if not re.fullmatch(r":?-{1,}:?", cell.strip()):
                return False
        return True

    tables: List[List[List[str]]] = []
    i = 0
    while i < n:
        line = lines[i]
        if "|" in line and i + 1 < n and is_separator(lines[i + 1]):
            header = split_row(line)
            width = len(header)
            rows = [header]
            j = i + 2
            while j < n and "|" in lines[j] and lines[j].strip() and not is_separator(lines[j]):
                row = split_row(lines[j])
                if len(row) < width:
                    row += [""] * (width - len(row))
                elif len(row) > width:
                    row = row[:width]
                rows.append(row)
                j += 1
            tables.append(rows)
            i = j
        else:
            i += 1
    return tables


# --------------------------------------------------------------------------- #
# From Markdown
# --------------------------------------------------------------------------- #
def markdown_to_html(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Rendering Markdown to HTML...")
    fragment = _md_to_html_fragment(_read_text(path))
    title = os.path.basename(os.path.splitext(path)[0])
    html_doc = (
        "<!DOCTYPE html>\n<html lang=\"en\">\n<head>\n<meta charset=\"utf-8\">\n"
        f"<title>{title}</title>\n<style>\n"
        "body{font-family:Segoe UI,Arial,sans-serif;max-width:820px;margin:2rem auto;"
        "padding:0 1rem;line-height:1.6;color:#222}"
        "table{border-collapse:collapse}td,th{border:1px solid #ccc;padding:6px 10px}"
        "code{background:#f4f4f4;padding:2px 4px;border-radius:3px}"
        "pre{background:#f4f4f4;padding:12px;border-radius:6px;overflow:auto}\n"
        "</style>\n</head>\n<body>\n" + fragment + "\n</body>\n</html>\n"
    )
    out = _unique_path(os.path.splitext(path)[0] + ".html")
    _write_text(out, html_doc)
    return [out]


def markdown_to_txt(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Converting Markdown to plain text...")
    fragment = _md_to_html_fragment(_read_text(path))
    extractor = _TextExtractor()
    extractor.feed(fragment)
    out = _unique_path(os.path.splitext(path)[0] + ".txt")
    _write_text(out, unescape(extractor.text()))
    return [out]


def markdown_to_docx(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Compiling Markdown into Word (.docx)...")
    from htmldocx import HtmlToDocx

    fragment = _md_to_html_fragment(_read_text(path))
    parser = HtmlToDocx()
    document = parser.parse_html_string(fragment)
    out = _unique_path(os.path.splitext(path)[0] + ".docx")
    document.save(out)
    return [out]


def markdown_to_xlsx(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Extracting tables into Excel (.xlsx)...")
    from openpyxl import Workbook

    md_text = _read_text(path)
    tables = extract_markdown_tables(md_text)
    workbook = Workbook()
    workbook.remove(workbook.active)

    if tables:
        for index, table in enumerate(tables, start=1):
            sheet = workbook.create_sheet(title=f"Table {index}"[:31])
            for row in table:
                sheet.append(row)
    else:
        sheet = workbook.create_sheet(title="Content")
        for line in md_text.splitlines():
            if line.strip():
                sheet.append([line.rstrip()])

    out = _unique_path(os.path.splitext(path)[0] + ".xlsx")
    workbook.save(out)
    return [out]


def markdown_to_csv(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Extracting tables into CSV...")
    md_text = _read_text(path)
    tables = extract_markdown_tables(md_text)
    base = os.path.splitext(path)[0]
    outputs: List[str] = []

    def write_csv(out_path: str, rows: List[List[str]]) -> None:
        with open(out_path, "w", encoding="utf-8", newline="") as handle:
            csv.writer(handle).writerows(rows)
        outputs.append(out_path)

    if not tables:
        rows = [[line.rstrip()] for line in md_text.splitlines() if line.strip()]
        write_csv(_unique_path(base + ".csv"), rows)
    elif len(tables) == 1:
        write_csv(_unique_path(base + ".csv"), tables[0])
    else:
        for index, table in enumerate(tables, start=1):
            write_csv(_unique_path(f"{base}_table{index}.csv"), table)
    return outputs


# ----- PDF rendering ------------------------------------------------------- #
def _register_pdf_font(pdf):
    """Register a Unicode font with bold/italic variants when available.

    Returns (family, markdown_ok). markdown_ok indicates bold/italic styles are
    available so fpdf2's markdown rendering can be used safely.
    """
    win = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts")
    bundled = resource_path("assets")
    font_sets = [
        (
            os.path.join(bundled, "DejaVuSans.ttf"),
            os.path.join(bundled, "DejaVuSans-Bold.ttf"),
            os.path.join(bundled, "DejaVuSans-Oblique.ttf"),
            os.path.join(bundled, "DejaVuSans-BoldOblique.ttf"),
        ),
        (
            os.path.join(win, "arial.ttf"),
            os.path.join(win, "arialbd.ttf"),
            os.path.join(win, "ariali.ttf"),
            os.path.join(win, "arialbi.ttf"),
        ),
    ]
    for regular, bold, italic, bold_italic in font_sets:
        if os.path.exists(regular):
            pdf.add_font("Uni", "", regular)
            has_bold = os.path.exists(bold)
            has_italic = os.path.exists(italic)
            if has_bold:
                pdf.add_font("Uni", "B", bold)
            if has_italic:
                pdf.add_font("Uni", "I", italic)
            if os.path.exists(bold_italic):
                pdf.add_font("Uni", "BI", bold_italic)
            return "Uni", (has_bold and has_italic)
    return "helvetica", True


def _render_markdown_pdf(md_text: str, pdf, family: str, markdown_ok: bool) -> None:
    from fpdf.enums import XPos, YPos

    def write_para(text: str, size: int = 11, bold: bool = False) -> None:
        style = "B" if (bold and markdown_ok) else ""
        pdf.set_font(family, style, size)
        content = text if markdown_ok else _strip_inline(text)
        pdf.multi_cell(
            0, size * 0.55 + 2, content,
            new_x=XPos.LMARGIN, new_y=YPos.NEXT, markdown=markdown_ok,
        )

    lines = md_text.splitlines()
    i = 0
    n = len(lines)
    para_buffer: List[str] = []

    def flush_para() -> None:
        if para_buffer:
            write_para(" ".join(para_buffer))
            pdf.ln(2)
            para_buffer.clear()

    while i < n:
        line = lines[i]
        stripped = line.strip()

        if "|" in line and i + 1 < n and re.fullmatch(
            r"\s*\|?[\s:|-]*-[\s:|-]*\|?\s*", lines[i + 1]
        ):
            flush_para()
            tables = extract_markdown_tables("\n".join(lines[i:]))
            if tables:
                table = tables[0]
                with pdf.table() as t:
                    for row in table:
                        tr = t.row()
                        for cell in row:
                            tr.cell(cell)
                pdf.ln(2)
                consumed = 2 + max(0, len(table) - 1)
                i += consumed
                continue

        if stripped.startswith("```"):
            flush_para()
            i += 1
            code_lines: List[str] = []
            while i < n and not lines[i].strip().startswith("```"):
                code_lines.append(lines[i])
                i += 1
            i += 1
            pdf.set_font(family, "", 9)
            pdf.multi_cell(
                0, 5, "\n".join(code_lines),
                new_x=XPos.LMARGIN, new_y=YPos.NEXT,
            )
            pdf.ln(2)
            continue

        heading = re.match(r"^(#{1,6})\s+(.*)$", stripped)
        if heading:
            flush_para()
            level = len(heading.group(1))
            write_para(heading.group(2), size=20 - level * 2, bold=True)
            pdf.ln(2)
            i += 1
            continue

        list_item = re.match(r"^\s*([-*+]|\d+\.)\s+(.*)$", line)
        if list_item:
            flush_para()
            bullet = "-" if list_item.group(1) in "-*+" else list_item.group(1)
            write_para(f"{bullet} {list_item.group(2)}")
            i += 1
            continue

        if re.fullmatch(r"(-{3,}|\*{3,}|_{3,})", stripped):
            flush_para()
            pdf.ln(3)
            i += 1
            continue

        if not stripped:
            flush_para()
            i += 1
            continue

        para_buffer.append(stripped)
        i += 1

    flush_para()


def markdown_to_pdf(path: str, status_cb: StatusCb = None) -> List[str]:
    _emit(status_cb, "Rendering Markdown to PDF...")
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    family, markdown_ok = _register_pdf_font(pdf)
    pdf.set_font(family, "", 11)
    _render_markdown_pdf(_read_text(path), pdf, family, markdown_ok)
    out = _unique_path(os.path.splitext(path)[0] + ".pdf")
    pdf.output(out)
    return [out]


_REVERSE_DISPATCH = {
    "docx": markdown_to_docx,
    "html": markdown_to_html,
    "pdf": markdown_to_pdf,
    "txt": markdown_to_txt,
    "xlsx": markdown_to_xlsx,
    "csv": markdown_to_csv,
}


def from_markdown(path: str, target: str, status_cb: StatusCb = None) -> List[str]:
    target = target.lower().lstrip(".")
    if target not in _REVERSE_DISPATCH:
        raise ValueError(f"Unsupported Markdown target: {target}")
    return _REVERSE_DISPATCH[target](path, status_cb)


def convert(
    path: str,
    status_cb: StatusCb = None,
    reverse_target: str = "docx",
    lang: str = "English",
    cancel_cb=None,
) -> List[str]:
    """High-level entry: route by extension. `.md` uses `reverse_target`;
    OCR paths use `lang`. `cancel_cb` is polled at checkpoints to abort early."""
    _check_cancel(cancel_cb)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".md", ".markdown"):
        return from_markdown(path, reverse_target, status_cb)
    return to_markdown(path, status_cb, lang, cancel_cb)
